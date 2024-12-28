from mip import *
import openpyxl
import string
import streamlit as st
import tempfile
from collections import defaultdict


def saitekika(file):
  #Excelファイル・シートを開く
  book = openpyxl.load_workbook(file)

  # sheethope = book['2023年度1月希望']
  # sheetkakutei = book['2023年度1月確定']
  sheethope = book['今月']
  sheetkakutei = book.create_sheet(title='今月確定')
  sheetsengetu = book['先月']
  
  #print(sheethope.cell(row=3,column=1).value)
  
  #定数用のデータの作成
  
  #従業員の名前のリスト
  ninzu = 0
  namae = []
  for cell in sheethope["A"]:
      #print (cell.value)
      if cell.value == "Ａ":
          break
      namae.append(cell.value)
      ninzu += 1
  ninzu -= 3
  namae = namae[3:]
  #print(ninzu)
  #print(namae)
  
  #今月の日数
  kongetunissu = 0
  for column in range(2,35):
      cell = sheethope.cell(row=3,column=column)
      #print (cell.value)
      if cell.value not in  {"月","火","水","木","金","土","日"}:
          break
      kongetunissu += 1
  #print("今月",kongetunissu)
  
  #先月の日数
  sengetunissu = 0
  for column in range(2,35):
      cell = sheetsengetu.cell(row=3,column=column)
      #print (cell.value)
      if cell.value not in  {"月","火","水","木","金","土","日"}:
          break
      sengetunissu += 1
  #print("先月",sengetunissu)
  
  
  
  kioku = {}
  
  I = namae#従業員
  D = [i+1 for i in range(kongetunissu)]#日数
  T = [i+1 for i in range(5)]#スロット(1:AB 2:CD 3:EF 4:ABCD 5:CDEF)
  
  #シフト希望読み込み
  k = {}#k={1:AB 2:CD 3:EF 4:ABCD 5:CDEF}
  syain = []
  sinjin = []
  sinjiniti = []
  syainkibou = {}
  sinjinkibou = {}
  hiku = defaultdict(int)
  for i in range(len(I)):
    n = I[i]
    if sheethope.cell(row=3+i+1 ,column=1).fill.fgColor.rgb == "FF92D050":
        sinjin.append(n)
        sinjiniti.append(4+i)
        for d in D:
          sinjinkibou[n,d]=sheethope.cell(row=3+i+1 ,column=1+d).value
          #print(n,d,sinjinkibou[n,d])
    if sheethope.cell(row=3+i+1 ,column=1).fill.fgColor.rgb == "FF0070C0":
        syain.append(n)
        for d in D:
          syainkibou[n,d]=sheethope.cell(row=3+i+1 ,column=1+d).value
          #print(n,d,syainkibou[n,d])
          if syainkibou[n,d] == None:
            continue
          if "AB" in syainkibou[n,d]:
            hiku[d,1] += 1
          if "CD" in syainkibou[n,d]:
            hiku[d,2] += 1
          if "EF" in syainkibou[n,d]:
            hiku[d,3] += 1
        continue
    for d in D:
      toridasi = sheethope.cell(row=3+i+1 ,column=1+d).value
      #print(n,d,toridasi)
      if toridasi == None:
        k[n,d,1] = 0
        k[n,d,2] = 0
        k[n,d,3] = 0
        k[n,d,4] = 0
        k[n,d,5] = 0
        continue
      sihutokibou = "".join([hanbetu for hanbetu in toridasi if hanbetu in string.ascii_uppercase or hanbetu =="2" or hanbetu =="4"])
      sonota = "".join([hanbetu for hanbetu in toridasi if not (hanbetu in string.ascii_uppercase or hanbetu =="2" or hanbetu =="4")])
      #print(n,d,sonota)
      if sonota != "":
        kioku[n,d] = (sonota,toridasi[0] in "ABCDEF")
      if sihutokibou == "ABCDEF":
        k[n,d,1] = 1
        k[n,d,2] = 1
        k[n,d,3] = 1
        k[n,d,4] = 1
        k[n,d,5] = 1
      if sihutokibou == "ABCDEF4":
        k[n,d,1] = 0
        k[n,d,2] = 0
        k[n,d,3] = 0
        k[n,d,4] = 1
        k[n,d,5] = 1
      if sihutokibou == "ABCDEF2":
        k[n,d,1] = 1
        k[n,d,2] = 1
        k[n,d,3] = 1
        k[n,d,4] = 0
        k[n,d,5] = 0
      elif sihutokibou == "ABCD":
        k[n,d,1] = 0
        k[n,d,2] = 0
        k[n,d,3] = 0
        k[n,d,4] = 1
        k[n,d,5] = 0
      elif sihutokibou == "CDEF":
        k[n,d,1] = 0
        k[n,d,2] = 0
        k[n,d,3] = 0
        k[n,d,4] = 0
        k[n,d,5] = 1
      elif sihutokibou == "ABCD2":
        k[n,d,1] = 1
        k[n,d,2] = 1
        k[n,d,3] = 0
        k[n,d,4] = 0
        k[n,d,5] = 0
      elif sihutokibou == "CDEF2":
        k[n,d,1] = 0
        k[n,d,2] = 1
        k[n,d,3] = 1
        k[n,d,4] = 0
        k[n,d,5] = 0
      elif sihutokibou == "ABEF2":
        k[n,d,1] = 1
        k[n,d,2] = 0
        k[n,d,3] = 1
        k[n,d,4] = 0
        k[n,d,5] = 0
      elif sihutokibou == "AB":
        k[n,d,1] = 1
        k[n,d,2] = 0
        k[n,d,3] = 0
        k[n,d,4] = 0
        k[n,d,5] = 0
      elif sihutokibou == "CD":
        k[n,d,1] = 0
        k[n,d,2] = 1
        k[n,d,3] = 0
        k[n,d,4] = 0
        k[n,d,5] = 0
      elif sihutokibou == "EF":
        k[n,d,1] = 0
        k[n,d,2] = 0
        k[n,d,3] = 1
        k[n,d,4] = 0
        k[n,d,5] = 0
      else:
        k[n,d,1] = 0
        k[n,d,2] = 0
        k[n,d,3] = 0
        k[n,d,4] = 0
        k[n,d,5] = 0
  
  #print(k)
  #print(kioku)
  #print(syain)
  #print(syainkibou)
  #print(hiku)
  
  #シフト結果出力用シートの追加
  for i,row in enumerate(sheethope.iter_rows()):
      for j,cell in enumerate(row):
        value = ""
        if not (i+1 >= 4 and i+1<=len(I) and j+1 >= 2 and j+1 <= kongetunissu+1 and not i+1 in sinjiniti):
          value = cell.value
        newcell=sheetkakutei.cell(row=cell.row, column=cell.column, value=value)
        if cell.has_style:
            newcell._style = cell._style
        #newcell.fill=openpyxl.styles.PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
  
  
  #社員かどうかの判断
  s = {}
  for i in range(len(I)):
    n=I[i]
    if sheethope.cell(row=3+i+1,column=1).fill.fgColor.rgb == "FF0070C0":#青色がついている日は休日
      s[n] = 1 #社員
    else:
      s[n] = 0 #非社員
  #print(s)
  
  #平日か休日の判断
  n = {}
  for d in D:
    if sheethope.cell(row=3 ,column=1+d).fill.fgColor.rgb == "FFFFC000":#オレンジ色がついている日は休日
      #print(d)
      n[d] = 1 #休日
    elif sheethope.cell(row=3 ,column=1+d).fill.fgColor.tint == -0.1499984740745262:#灰色の日は閉館
      n[d] = 2 #閉館
    else:
      n[d] = 0
  #print(sheethope.cell(row=3 ,column=7).fill.fgColor.rgb == "FFFFC000")
  #print(n)
  
  #n={1:2,2:2,3:2,4:0,5:0,6:1,7:1,8:1,9:0,10:0,11:0,12:0,13:1,14:1,15:0,16:0,17:0,18:0,19:0,20:1,21:1,22:0,23:0,24:0,25:0,26:0,27:1,28:1,29:0,30:0,31:2}
  
  # for i in range(31):
  #   if n[i+1] == 2:
  #     print(i+1,"＝休館")
  #   elif n[i+1] == 1:
  #     print(i+1,"=休日")
  #   else:
  #     print(i+1,"=平日")
  
  #平日か休日か判断してシフトに必要な人数を出す
  p = {}
  for n_d in n:
    #print(n[n_d])
    if n[n_d] == 1:  #休日
      p[n_d,1] = 6 - hiku[n_d,1]
      p[n_d,2] = 6 - hiku[n_d,2]
      p[n_d,3] = 2 - hiku[n_d,3]
      for t in range(1,4):
        if p[n_d,t] == 0:
          p[n_d,t] = 1
    elif n[n_d] == 2:#閉館
      p[n_d,1] = 0
      p[n_d,2] = 0
      p[n_d,3] = 0
    else:            #平日
      p[n_d,1] = 5 - hiku[n_d,1]
      p[n_d,2] = 5 - hiku[n_d,2]
      p[n_d,3] = 2 - hiku[n_d,3]
      for t in range(1,4):
        if p[n_d,t] == 0:
          p[n_d,t] = 1
  #print(p)
  
  #社員の各スロット必要人数
  # sp = {}
  # for n_d in n:
  #   if n[n_d] == 1:  #休日
  #     sp[n_d,1] = 2
  #     sp[n_d,2] = 2
  #     sp[n_d,3] = 1
  #   elif n[n_d] == 2:#閉館
  #     sp[n_d,1] = 0
  #     sp[n_d,2] = 0
  #     sp[n_d,3] = 0
  #   else:            #平日
  #     sp[n_d,1] = 2
  #     sp[n_d,2] = 2
  #     sp[n_d,3] = 1
  
  #各スロットの実働時間
  a = {}
  a[1] = 4.5
  a[2] = 4.5
  a[3] = 4
  a[4] = 8
  a[5] = 7.5
  
  #従業員の出勤希望回数
  kai={}
  for i in range(len(I)):
    n=I[i]
    kai[n] = sheethope.cell(row=3+i+1 ,column=kongetunissu+3).value
  
  #kai={1:None,2:11,3:None,4:12,5:15,6:13,7:16,8:15,9:14,10:None,11:15,12:None,13:None,14:12,15:16,16:None,17:None,18:None,19:None,20:12,21:15,22:None,23:None,24:None,25:None,26:None,27:None,28:None,29:10,30:10,31:10,32:10,33:None}
  #print(kai)
  
  #主婦さんの点数(月の1日～10日は講座の受付があり、申し込み受付に慣れている主婦さんに多く入ってもらうため)
  syuhu={}
  for i in range(len(I)):
    n=I[i]
    if sheethope.cell(row=3+i+1 ,column=1).fill.fgColor.rgb == "FFFFFF00":#黄色なら主婦さん
      syuhu[n] = 1
    else:
      syuhu[n] = 0
  #print(syuhu)
  
  syuhuten = {}
  for i in I:
    for d in D:
      if syuhu[i] == 1:
        if d <=10:
            syuhuten[i,d,1] = 10
            syuhuten[i,d,2] = 10
            syuhuten[i,d,3] = 0
            syuhuten[i,d,4] = 10
            syuhuten[i,d,5] = 0
        else:
            syuhuten[i,d,1] = 0
            syuhuten[i,d,2] = 0
            syuhuten[i,d,3] = 0
            syuhuten[i,d,4] = 0
            syuhuten[i,d,5] = 0
      else:
        for t in T:
          syuhuten[i,d,t] = 0
  #print(syuhuten)
  
  
  
  
  
  
  
  #空問題の作成
  model = Model('Shift')
  
  
  
  
  
  
  
  
  #決定変数の作成
  x = {}
  for i in I:
      if i in syain:
        break
      for d in D:
          for t in T:
              x[i, d, t] = model.add_var(f'x{i},{d},{t}', var_type='B')
  
  #シフト希望が必要人数に足りない場合のペナルティ変数(最小化したい)
  y = {}
  for d in D:
      for t in T:
        y[d,t] = model.add_var(f'y{d},{t}', var_type='I')
  
  #社員がAB・CD・EFに2人ずつ入れないときに、+1するようのペナルティ変数
  #(最小化したい、+1なのは、必ず1人は入らないといけないから)
  # e = {}
  # for d in D:
  #   for t in T:
  #     e[d,t] = model.add_var(f'e{d},{t}', var_type='I')
  
  #希望出勤回数のためのペナルティ変数(最小化したい)
  h = {}
  for i in I:
    h[i] = model.add_var(f'h{i}', var_type='I')
  
  #連勤を確認するための決定変数
  z = {}
  for i in I:
      for d in range(1,len(D)+6):
        z[i,d] = model.add_var(f'z{i},{d}', var_type='B')
  
  
  
  
  
  
  
  
  #制約条件の追加
  #勤務可能なら割り当てる
  for i in I:
    if i in syain:
        break
    for d in D:
      for t in T:
        model += x[i,d,t]<=k[i,d,t]
  
  #その時間帯に必要な人数を割り当てる
  for d in D:
      model += xsum(x[i,d,1] + x[i,d,4] for i in I if i not in syain) == p[d,1] - y[d,1]
      model += xsum(x[i,d,2] + x[i,d,4] + x[i,d,5] for i in I if i not in syain) == p[d,2] - y[d,2]
      model += xsum(x[i,d,3] + x[i,d,5] for i in I if i not in syain) == p[d,3] - y[d,3] 
  
  #AB・CD・EFの各時間帯にできれば2人、最低1人社員を出勤させる
  # for d in D:
  #       model += xsum(x[i,d,1]+x[i,d,4] for i in I if s[i] == 1) == sp[d,1] - e[d,1]
  #       model += xsum(x[i,d,2]+x[i,d,4]+x[i,d,5] for i in I if s[i] == 1) == sp[d,2] - e[d,2]
  #       model += xsum(x[i,d,3]+x[i,d,5] for i in I if s[i] == 1) == sp[d,3] - e[d,3]
  
  #1日の実働時間条件
  for i in I:
    if i in syain:
      break
    for d in D:
      model += xsum(x[i,d,t] for t in T) <= 1
  
  #1か月の総実働時間条件
  for i in I:
    if s[i] == 0:#アルバイトのとき
      model += xsum(a[t]*x[i,d,t] for d in D for t in T) <= 80
  
  #1か月の出勤回数の希望
  for i in I:
    if kai[i] != None and s[i] == 0:
      model += xsum(x[i,d,t] for d in D for t in T) - kai[i] <= h[i]
      model += xsum(x[i,d,t] for d in D for t in T) - kai[i] >= -h[i]
  
  #前月の連勤データの取得
  for i in range(len(I)):
    n=I[i]
    for d in range(1,6):
      #print(sheetsengetu.cell(row=2 ,column=sengetunissu - 5 + 1 + d).value)
      if sheetsengetu.cell(row=4+i ,column=sengetunissu - 5 + 1 + d).value != None and sheetsengetu.cell(row=4+i ,column=sengetunissu - 5 + 1 + d).value != "×" and sheetsengetu.cell(row=4+i ,column=sengetunissu - 5 + 1 + d).value != "/":
        z[n,d] = 1
  
  #今月の連勤データの取得
  for i in namae:
    if i in syain:
      break
    for d in D:
      z[i,d+5] = xsum(x[i,d,t] for t in T)
  
  #連勤に関する制約条件(5日までならOK、6日以上はだめ)
  for i in namae:
    for d in D:
      model += z[i,d] + z[i,d+1] + z[i,d+2] + z[i,d+3] + z[i,d+4] + z[i,d+5] <= 5
  
  
  
  
  
  
  
  
  
  #目的関数の設定
  model.objective = minimize(1000*xsum(y[d,t] for i in I for d in D for t in T)+0.1*xsum(h[i] for i in I if i not in syain)-xsum(syuhuten[i,d,t]*x[i,d,t] for i in I for d in D for t in T if i not in syain))
  
  #LP形式
  model.write('Shift.lp')
  
  
  
  
  
  
  
  
  #最適化の実行
  status = model.optimize()
  
  #最適化の結果出力
  if status == OptimizationStatus.OPTIMAL:
      #print('最適値 =', model.objective_value)
      print('y =', model.objective_value)
      T_Name = {1:'AB', 2:'CD', 3:'EF',4:'ABCD',5:'CDEF'}
      for i in range(len(I)):
          n=I[i]
          if n in syain:
            for d in D:
              sheetkakutei.cell(row=3+i+1, column=1+d).value = syainkibou[n,d]
            continue
          for d in D:
              txt=""
              for t in T:
                  if x[n,d,t].x > 0.01:
                      txt+=T_Name[t]
              if (n,d) in kioku:
                if kioku[n,d][1]:
                  txt += kioku[n,d][0]
                else:
                  txt = kioku[n,d][0] + txt
              if txt != "":
                  sheetkakutei.cell(row=3+i+1, column=1+d).value = txt    #スロットを出力
      book.save("シフト結果.xlsx")
  
      #return "シフト結果.xlsx"
      #print(temp_file.name
      #for i in I:
        #print(i,x[i,4,1].x)
  else:
      print('最適解が求まりませんでした。')

  return("シフト結果.xlsx")









def main():
  st.title("シフト自動作成")
  file=st.file_uploader("アップロードしてください",type="xlsx")
  if file is not None:
    name=saitekika(file)
    print(open)
    with open(name,"rb") as f:
      st.download_button(
          label="結果をダウンロード",
          data=f,
          file_name="シフト結果.xlsx",
          mime="application/vnd.ms-excel"
      )
main()
